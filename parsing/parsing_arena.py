import shutil
import time
import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from config.config import parsing_url_arena, source_path_arena, path_to_download, project_path, destination_path


def parser_tms_arena():
    driver = webdriver.Chrome()
    driver.get(parsing_url_arena)  # Подключаемся на TMS ArenaCity
    driver.maximize_window()
    time.sleep(3)
    element = driver.find_element(By.XPATH,
                                  '//*[@id="tms-view-body"]/div/div/div[3]/div[2]/div[2]/form/div[1]/input[2]')
    element.send_keys("Limon4ik")
    driver.find_element(By.XPATH, '//*[@id="tms-view-body"]/div/div/div[3]/div[2]/div[2]/form/div[3]/button').click()
    time.sleep(1)
    driver.find_element(By.XPATH,
                        '/html/body/tms/div/div[5]/header/div[5]/tms-navigation-header-menu/div[4]/div/div[1]').click()
    time.sleep(1)
    driver.find_element(By.XPATH,
                        '//*[@id="tms-view-body"]/tms-schedule-builder/div/div/tms-date-range-picker/div/button[2]').click()
    time.sleep(1)
    driver.find_element(By.XPATH,
                        '//*[@id="tms-view-body"]/tms-schedule-builder/div/div/tms-export-button/tms-dropdown-actions/tms-dropdown/button/i').click()
    time.sleep(0.5)
    driver.find_element(By.XPATH,
                        '//*[@id="tms-view-body"]/tms-schedule-builder/div/div/tms-export-button/tms-dropdown-actions/tms-dropdown/div/div/tms-list-select/div[1]/tms-list-select-item[3]/div/span').click()
    time.sleep(1)
    driver.close()
    driver.quit()
    os.chdir(path_to_download)

    if os.path.exists(source_path_arena):
        shutil.move(source_path_arena, destination_path)
    print(f'{"*" * 10} Файл Schedule - BY_SS_ArenaCity.xlsx перемещен в data/ {"*" * 10}')
    time.sleep(2)

    wookbook = openpyxl.load_workbook(f'{project_path}\data\Schedule - BY_SS_ArenaCity.xlsx')
    worksheet = wookbook.active
    data = []
    for i in range(1, worksheet.max_row):
        temp = []
        for col in worksheet.iter_cols(1, worksheet.max_column - 2):
            temp.append(col[i].value)
        del temp[2], temp[2], temp[2], temp[2], temp[2]
        data.append(tuple(temp))
    print(f'{"*" * 10} Загрузка данных из TMS ARENA прошла успешно! {"*" * 10}')
    return data
