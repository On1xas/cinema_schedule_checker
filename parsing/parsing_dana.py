import shutil
import time
import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from config.config import parsing_url_dana, source_path_dana, project_path, destination_path, path_to_download
import sqlite3
def parser_tms_dana():
    driver = webdriver.Chrome()
    driver.get(parsing_url_dana)  # Подключаемся на TMS Dana
    driver.maximize_window()
    time.sleep(3)
    # Ставим курсор в поле ввода пароля
    element = driver.find_element(By.XPATH, '//*[@id="tms-view-body"]/div/div/div[2]/div[2]/div[2]/form/div[1]/input[2]')
    # Вводим пароль
    element.send_keys("Limon4ik")
    # Нажимаем кнопку Login
    driver.find_element(By.XPATH, '//*[@id="tms-view-body"]/div/div/div[2]/div[2]/div[2]/form/div[3]/button').click()
    time.sleep(1)
    # Нажимаем Schedule
    driver.find_element(By.XPATH,
                        '/html/body/tms/div/div[5]/header/div[5]/tms-navigation-header-menu/div[4]/div/div[1]').click()
    time.sleep(1)
    # Нажимаем Прогрузить неделю
    driver.find_element(By.XPATH,
                        '//*[@id="tms-view-body"]/tms-schedule-builder/div/div/tms-date-range-picker/div/button[2]').click()
    time.sleep(1)
    driver.find_element(By.XPATH,
                        '//*[@id="tms-view-body"]/tms-schedule-builder/div/div/tms-export-button/tms-dropdown-actions/tms-dropdown/button/i').click()
    time.sleep(0.5)
    driver.find_element(By.XPATH,
                        '//*[@id="tms-view-body"]/tms-schedule-builder/div/div/tms-export-button/tms-dropdown-actions/tms-dropdown/div/div/tms-list-select/div[1]/tms-list-select-item[3]/div/span').click()
    time.sleep(1)
    driver.close()
    driver.quit()
    os.chdir(path_to_download)
    if os.path.exists(source_path_dana):
        shutil.move(source_path_dana, destination_path)
        print(f'{"*" * 10} Файл Schedule - BY_SS_Dana.xlsx перемещен в data/ {"*" * 10}')
    time.sleep(2)
    wookbook = openpyxl.load_workbook(f'{project_path}\data\Schedule - BY_SS_Dana.xlsx')
    worksheet = wookbook.active
    data = []
    for i in range(1, worksheet.max_row):
        temp = []
        for col in worksheet.iter_cols(1, worksheet.max_column - 2):
            temp.append(col[i].value)
        del temp[2], temp[2], temp[2], temp[2], temp[2]
        data.append(tuple(temp))
    print(f'{"*" * 10} Загрузка данных из TMS DANA прошла успешно! {"*" * 10}')
    return data



